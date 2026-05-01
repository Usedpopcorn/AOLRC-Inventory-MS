from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app import create_app, db  # noqa: E402
from app.models import (  # noqa: E402
    Check,
    CheckLine,
    CountLine,
    CountSession,
    InventoryAdminEvent,
    Item,
    OrderBatch,
    OrderLine,
    SupplyNote,
    Venue,
    VenueItem,
    VenueItemCount,
    VenueNote,
)
from app.services.inventory_rules import (  # noqa: E402
    exact_item_name_key,
    find_similar_items,
)
from app.services.inventory_status import suggest_status_from_count  # noqa: E402

INFO_SHEET = "Info"
ORDER_SHEETS = {"Quarterly Orders", "Monthly Orders"}
VENUE_NAME_ALIASES = {
    "Ananda": "Ananda Hall",
    "Gita": "Gita Hall",
    "Shakti": "Shakti Hall",
    "MMH": "Main Meditation Hall",
}
ITEM_NAME_ALIASES = {
    "Back Jacks": "Backjacks",
    "Meditation Bolsters": "Bolsters",
    "Meditation Cushions": "Cushions",
    "Notebooks": "Notepads",
    "Copy Paper (packs)": "Copy Paper",
    "Feedback Forms (purple)": "Feedback Forms",
    "Feedback Forms (QR)": "QR Code Cards",
    "Markers (coloring)": "Dry Erase Markers",
    "Markers": "Dry Erase Markers",
}
GROUP_ITEM_TYPE_DEFAULTS = {
    "01": "durable",
    "02": "consumable",
    "03": "consumable",
    "04": "durable",
    "05": "durable",
}
RESET_MODELS = [
    CheckLine,
    Check,
    CountLine,
    CountSession,
    VenueItemCount,
    VenueNote,
    SupplyNote,
    InventoryAdminEvent,
    OrderLine,
    OrderBatch,
    VenueItem,
]


@dataclass(frozen=True)
class SpreadsheetRow:
    venue_name: str
    group_code: str | None
    group_label: str | None
    item_name: str
    par_value: int | None
    count_value: int | None
    status_text: str | None
    notes: str | None


def _collapse_spaces(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _to_int_or_none(value: object) -> int | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _normalize_group_code(value: object) -> str | None:
    text = _collapse_spaces(value)
    if not text:
        return None
    match = re.search(r"\d+", text)
    if not match:
        return None
    return match.group(0).zfill(2)


def _parse_group_labels(workbook) -> dict[str, str]:
    labels: dict[str, str] = {}
    if INFO_SHEET not in workbook.sheetnames:
        return labels
    worksheet = workbook[INFO_SHEET]
    for row in worksheet.iter_rows(min_row=1, values_only=True):
        first_cell = _collapse_spaces(row[0] if row else "")
        if not first_cell:
            continue
        match = re.match(r"^(\d{1,2})\s*-\s*(.+)$", first_cell)
        if not match:
            continue
        labels[match.group(1).zfill(2)] = _collapse_spaces(match.group(2))
    return labels


def _resolve_venue_name(raw_name: str) -> str:
    return VENUE_NAME_ALIASES.get(raw_name, raw_name)


def _parse_rows(workbook) -> list[SpreadsheetRow]:
    group_labels = _parse_group_labels(workbook)
    rows: list[SpreadsheetRow] = []
    sheet_names = [s for s in workbook.sheetnames if s not in ({INFO_SHEET} | ORDER_SHEETS)]
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        resolved_venue_name = _resolve_venue_name(_collapse_spaces(sheet_name))
        for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
            group_code = _normalize_group_code(raw_row[0])
            raw_item_name = _collapse_spaces(raw_row[1])
            if not raw_item_name:
                continue
            rows.append(
                SpreadsheetRow(
                    venue_name=resolved_venue_name,
                    group_code=group_code,
                    group_label=group_labels.get(group_code or ""),
                    item_name=raw_item_name,
                    count_value=_to_int_or_none(raw_row[2]),
                    par_value=_to_int_or_none(raw_row[4]),
                    status_text=_collapse_spaces(raw_row[5]) or None,
                    notes=_collapse_spaces(raw_row[6]) or None,
                )
            )
    return rows


def _infer_item_type(group_code: str | None) -> str:
    return GROUP_ITEM_TYPE_DEFAULTS.get(group_code or "", "consumable")


def _resolve_or_create_item(row: SpreadsheetRow, stats: dict[str, int]):
    existing_by_exact = {
        exact_item_name_key(item.name): item for item in Item.query.order_by(Item.id.asc()).all()
    }
    desired_name = ITEM_NAME_ALIASES.get(row.item_name, row.item_name)
    desired_exact = exact_item_name_key(desired_name)
    item = existing_by_exact.get(desired_exact)
    matched_by = "alias_or_exact"
    if item is None:
        similar = find_similar_items(desired_name, limit=1)
        if similar and (
            similar[0].get("score", 0) >= 0.86
            or (similar[0].get("contains_match") and similar[0].get("score", 0) >= 0.78)
        ):
            item = Item.query.get(similar[0]["id"])
            matched_by = "similar"
    if item is None:
        inferred_type = _infer_item_type(row.group_code)
        item = Item(
            name=desired_name,
            item_type=inferred_type,
            item_category=inferred_type,
            tracking_mode="quantity",
            active=True,
            setup_group_code=row.group_code,
            setup_group_label=row.group_label,
        )
        db.session.add(item)
        db.session.flush()
        stats["created_items"] += 1
        return item

    item.active = True
    if row.group_code:
        item.setup_group_code = row.group_code
    if row.group_label:
        item.setup_group_label = row.group_label
    if matched_by == "similar":
        stats["matched_similar_items"] += 1
    else:
        stats["matched_existing_items"] += 1
    return item


def _resolve_or_create_venue(venue_name: str, stats: dict[str, int]):
    existing = Venue.query.filter(db.func.lower(Venue.name) == venue_name.lower()).first()
    if existing is not None:
        existing.active = True
        stats["matched_existing_venues"] += 1
        return existing
    venue = Venue(name=venue_name, active=True)
    db.session.add(venue)
    db.session.flush()
    stats["created_venues"] += 1
    return venue


def _reset_operational_tables(stats: dict[str, int]):
    for model in RESET_MODELS:
        deleted = model.query.delete(synchronize_session=False)
        stats[f"reset_{model.__tablename__}"] = int(deleted or 0)


def _sync_primary_key_sequence(table_name: str):
    sequence_name = f"{table_name}_id_seq"
    db.session.execute(
        db.text(
            "SELECT setval(:seq_name, COALESCE((SELECT MAX(id) FROM "
            + table_name
            + "), 1), true)"
        ),
        {"seq_name": sequence_name},
    )


def _apply_import(rows: list[SpreadsheetRow], dry_run: bool):
    stats: dict[str, int] = defaultdict(int)
    pair_data: dict[tuple[int, int], SpreadsheetRow] = {}
    count_rows_by_pair: dict[tuple[int, int], int] = {}
    status_rows_by_pair: dict[tuple[int, int], str] = {}

    _sync_primary_key_sequence("items")
    _sync_primary_key_sequence("venues")

    for row in rows:
        venue = _resolve_or_create_venue(row.venue_name, stats)
        item = _resolve_or_create_item(row, stats)
        pair_key = (venue.id, item.id)
        pair_data[pair_key] = row

        if row.count_value is not None:
            count_rows_by_pair[pair_key] = row.count_value
            stats["rows_with_count"] += 1
        if row.count_value is not None and row.par_value is not None and row.par_value > 0:
            derived_status = suggest_status_from_count(
                raw_count=row.count_value,
                par_value=row.par_value,
                tracking_mode="quantity",
            )
            if derived_status:
                status_rows_by_pair[pair_key] = derived_status
                stats["rows_with_derived_status"] += 1
        else:
            stats["rows_default_not_checked"] += 1

    if dry_run:
        stats["unique_venue_item_pairs"] = len(pair_data)
        db.session.flush()
        return stats

    _reset_operational_tables(stats)
    for table_name in (
        "venue_items",
        "checks",
        "check_lines",
        "count_sessions",
        "count_lines",
        "venue_item_counts",
    ):
        _sync_primary_key_sequence(table_name)

    venue_item_rows = []
    for (venue_id, item_id), row in pair_data.items():
        venue_item_rows.append(
            VenueItem(
                venue_id=venue_id,
                item_id=item_id,
                active=True,
                expected_qty=row.par_value,
                reorder_threshold=None,
            )
        )
    if venue_item_rows:
        db.session.bulk_save_objects(venue_item_rows)
    stats["venue_item_rows"] = len(venue_item_rows)

    check_by_venue: dict[int, Check] = {}
    count_session_by_venue: dict[int, CountSession] = {}

    for (venue_id, item_id), status_value in status_rows_by_pair.items():
        if venue_id not in check_by_venue:
            check = Check(venue_id=venue_id, user_id=None)
            db.session.add(check)
            db.session.flush()
            check_by_venue[venue_id] = check
        db.session.add(
            CheckLine(
                check_id=check_by_venue[venue_id].id,
                item_id=item_id,
                status=status_value,
            )
        )
    stats["check_sessions"] = len(check_by_venue)
    stats["check_lines"] = len(status_rows_by_pair)

    for (venue_id, item_id), count_value in count_rows_by_pair.items():
        if venue_id not in count_session_by_venue:
            session = CountSession(venue_id=venue_id, user_id=None)
            db.session.add(session)
            db.session.flush()
            count_session_by_venue[venue_id] = session
        db.session.add(
            CountLine(
                count_session_id=count_session_by_venue[venue_id].id,
                item_id=item_id,
                raw_count=count_value,
            )
        )
        db.session.add(
            VenueItemCount(
                venue_id=venue_id,
                item_id=item_id,
                raw_count=count_value,
            )
        )
    stats["count_sessions"] = len(count_session_by_venue)
    stats["count_lines"] = len(count_rows_by_pair)
    stats["venue_item_counts"] = len(count_rows_by_pair)
    stats["unique_venue_item_pairs"] = len(pair_data)

    db.session.commit()
    return stats


def _print_summary(stats: dict[str, int], *, dry_run: bool, workbook_path: Path, row_count: int):
    mode = "DRY RUN" if dry_run else "APPLY"
    print("")
    print(f"[{mode}] Spreadsheet import summary")
    print(f"Workbook: {workbook_path}")
    print(f"Rows parsed: {row_count}")
    for key in sorted(stats):
        print(f"{key}: {stats[key]}")
    print("")


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Reset operational inventory timeline and import data "
            "from venue inventory workbook."
        )
    )
    parser.add_argument(
        "--workbook",
        default=r"c:\Users\Jacob\Downloads\Copy of Venue Inventory .xlsx",
        help="Path to the source .xlsx workbook.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and compute import mapping without mutating database data.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook file not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    rows = _parse_rows(workbook)
    if not rows:
        raise SystemExit("No importable rows were parsed from workbook.")

    app = create_app()
    with app.app_context():
        stats = _apply_import(rows=rows, dry_run=bool(args.dry_run))
        if args.dry_run:
            db.session.rollback()
        _print_summary(
            stats,
            dry_run=bool(args.dry_run),
            workbook_path=workbook_path,
            row_count=len(rows),
        )


if __name__ == "__main__":
    main()
